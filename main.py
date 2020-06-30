from openpyxl import load_workbook, Workbook
from datetime import datetime

# 要求平均值的列的列表
avg_column_index_list = [2, 3, 4, 6]
# 时间戳列数
timestamp_column = 1
# 数据行有几列
column_count = 0

# 读取Workbook
workbook = load_workbook('sample.xlsx')
# 要操作的sheet
sheet = workbook.active
new_workbook = Workbook()
new_sheet = new_workbook.active
# 遍历的行列表
row_list = sheet.iter_rows(values_only=True)
# 下一行是否是数据行标记
next_row_is_data_row = False

sum_list = []
for x in avg_column_index_list:
    sum_list.append(0)
data_row_count = 0

for i, row in enumerate(row_list):
    if row[0] == '[root]' and data_row_count != 0:
        avg_row = []
        for column in range(column_count):
            if column not in avg_column_index_list:
                avg_row.append(None)
            else:
                k = avg_column_index_list.index(column)
                s = sum_list[k]
                avg = s / data_row_count
                avg_row.append(avg)
        new_sheet.append(avg_row)
        new_sheet.append(())
        sum_list = [0 for s in sum_list]
        data_row_count = 0

    # 判断这行是否是数据行
    if next_row_is_data_row:
        data_row_count = data_row_count + 1

        for j, column in enumerate(avg_column_index_list):
            sum_list[j] = sum_list[j] + float(row[column])

        new_row = list(row)
        new_row[timestamp_column] = datetime.fromtimestamp(new_row[timestamp_column])
        new_sheet.append(new_row)
    else:
        new_sheet.append(row)

    # 判断下一行是否是数据行
    if row[0] == '#':
        next_row_is_data_row = True
        column_count = len(row)
    else:
        next_row_is_data_row = False

avg_row = []
for column in range(column_count):
    if column not in avg_column_index_list:
        avg_row.append(None)
    else:
        k = avg_column_index_list.index(column)
        s = sum_list[k]
        avg = s / data_row_count
        avg_row.append(avg)
new_sheet.append(())
new_sheet.append(avg_row)

new_workbook.save('result.xlsx')
